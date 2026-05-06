"""
高校财务报销模拟数据生成脚本
- 纳入四大风险类型：源头虚构、金额操控、支出失控、流程违规
- 修复风险分布不平衡：高/中/低风险均匀分布于全部6类
- 新增字段：风险类型标签、具体异常行为描述
"""
import random
import string
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(2024)

# ─── 基础数据池 ──────────────────────────────────────────────
DEPARTMENTS = [
    "教务处", "科研处", "学生处", "后勤处", "财务处", "图书馆",
    "计算机学院", "经济管理学院", "理工学院", "医学院", "外国语学院",
    "艺术学院", "体育学院", "马克思主义学院", "继续教育学院", "研究生院"
]
APPLICANTS = [
    ("张伟", "教授"), ("李敏", "副教授"), ("王芳", "讲师"), ("刘洋", "助教"),
    ("陈磊", "行政专员"), ("赵丽", "科研助理"), ("孙强", "教授"), ("周婷", "副教授"),
    ("吴杰", "讲师"), ("郑华", "行政主任"), ("钱鑫", "教授"), ("马云飞", "讲师"),
    ("高峰", "副教授"), ("林晓", "助理研究员"), ("许明", "行政专员"),
    ("何静", "教授"), ("谢宇", "讲师"), ("宋阳", "副教授"), ("唐志远", "教授"),
    ("冯燕", "行政专员")
]
APPROVERS = ["王副校长", "李院长", "张处长", "刘副处长", "陈主任"]
PROJECTS = [
    "横向科研项目-2024001", "纵向科研项目-国自科基金", "教学改革项目-2024",
    "学科建设经费", "人才引进专项", "教育部重点项目", "省级科研基金",
    "校级科研启动基金", "国家社科基金", "教育厅创新项目"
]
BANKS = ["工商银行", "建设银行", "农业银行", "中国银行", "招商银行", "邮储银行"]

def rand_date(start="2024-01-01", end="2024-12-31"):
    s = datetime.strptime(start, "%Y-%m-%d")
    e = datetime.strptime(end, "%Y-%m-%d")
    return s + timedelta(days=random.randint(0, (e - s).days))

def rand_id(prefix="BX", n=8):
    return prefix + "".join(random.choices(string.digits, k=n))

def rand_person():
    return random.choice(APPLICANTS)

def rand_dept():
    return random.choice(DEPARTMENTS)

# ─── 风险注入引擎 ────────────────────────────────────────────
# 四大风险类型的具体表现清单
RISK_PATTERNS = {
    "源头虚构": [
        ("使用虚假发票套取资金", 6, "高风险"),
        ("真实发票重复报销", 5, "高风险"),
        ("个人消费以公务名义报销", 5, "高风险"),
        ("电子发票二次报销", 5, "高风险"),
        ("虚构差旅出行记录", 6, "高风险"),
        ("冒用他人身份领取劳务费", 6, "高风险"),
        ("虚构会议事项套取经费", 5, "高风险"),
        ("签订虚假咨询服务合同", 6, "高风险"),
        ("伪造设备采购合同套取经费", 6, "高风险"),
        ("围标、关联交易虚假采购", 5, "高风险"),
    ],
    "金额操控": [
        ("单笔大额拆分为多笔规避审批", 4, "中风险"),
        ("拆单规避集中采购规定", 4, "中风险"),
        ("将经费转入关联公司实现体外循环", 6, "高风险"),
        ("以合作研究名义违规转拨经费", 5, "高风险"),
        ("发票金额与实际支付不一致（优惠券/红包）", 3, "中风险"),
        ("虚报金额高于实际消费", 4, "中风险"),
        ("多人分散报销同一项目费用", 3, "中风险"),
    ],
    "支出失控": [
        ("专项经费挪用于日常接待", 4, "中风险"),
        ("科研经费中夹报个人消费", 4, "中风险"),
        ("课题经费用于非指定用途", 4, "中风险"),
        ("差旅期间夹报私人车辆费用", 3, "中风险"),
        ("向项目无关人员发放劳务费", 4, "中风险"),
        ("向家属发放劳务报酬", 5, "高风险"),
        ("给项目组成员违规发放专家咨询费", 4, "中风险"),
        ("超标准报销住宿、餐饮费用", 3, "中风险"),
        ("报销与业务明显不符的高额费用", 4, "中风险"),
    ],
    "流程违规": [
        ("先采购后补申请，流程倒置", 3, "中风险"),
        ("未通过合规渠道绕过集中采购", 3, "中风险"),
        ("合同签订晚于实际服务交付", 3, "中风险"),
        ("服务商不具资质却出具发票", 4, "中风险"),
        ("随意变更预算用途未履行审批", 3, "中风险"),
        ("差旅人员出行情况缺乏核实", 2, "中风险"),
        ("仅凭名单确认劳务有效性", 2, "中风险"),
        ("补签审批情况多发", 2, "中风险"),
        ("审批人与报销人为同一人", 4, "中风险"),
        ("不相容岗位未分离", 3, "中风险"),
    ],
    "正常": [
        ("正常报销", 0, "低风险"),
        ("正常报销", 0, "低风险"),
        ("正常报销", 1, "低风险"),
        ("正常报销", 0, "低风险"),
    ]
}

# 每个类别中高/中/低风险的目标比例
RISK_DISTRIBUTION = {
    # (低风险比, 中风险比, 高风险比)  — 必须加总=1.0
    "日常报销":   (0.55, 0.30, 0.15),
    "差旅费":     (0.50, 0.33, 0.17),
    "资产采购":   (0.45, 0.35, 0.20),
    "借款暂付款": (0.40, 0.38, 0.22),
    "劳务费":     (0.45, 0.33, 0.22),
    "其他费用":   (0.50, 0.30, 0.20),
}

def pick_risk_pattern(category):
    """根据类别目标分布，随机选取风险模式"""
    low_r, mid_r, high_r = RISK_DISTRIBUTION.get(category, (0.5, 0.3, 0.2))
    roll = random.random()
    if roll < low_r:
        return "正常", random.choice(RISK_PATTERNS["正常"])
    elif roll < low_r + mid_r:
        # 中风险：金额操控 + 支出失控 + 流程违规 各占一部分
        risk_type = random.choices(
            ["金额操控", "支出失控", "流程违规"],
            weights=[0.3, 0.4, 0.3]
        )[0]
        # 只取中风险条目
        candidates = [p for p in RISK_PATTERNS[risk_type] if p[2] == "中风险"]
        return risk_type, random.choice(candidates)
    else:
        # 高风险：源头虚构 + 金额操控高风险 + 支出失控高风险
        risk_type = random.choices(
            ["源头虚构", "金额操控", "支出失控"],
            weights=[0.5, 0.25, 0.25]
        )[0]
        candidates = [p for p in RISK_PATTERNS[risk_type] if p[2] == "高风险"]
        return risk_type, random.choice(candidates)

def derive_status(risk_level):
    if risk_level == "高风险":
        return random.choices(
            ["已拒绝", "待审计", "冻结待查"],
            weights=[0.4, 0.4, 0.2]
        )[0]
    elif risk_level == "中风险":
        return random.choices(
            ["待人工审核", "退回修改", "已通过", "挂起待核实"],
            weights=[0.35, 0.25, 0.30, 0.10]
        )[0]
    else:
        return random.choices(
            ["RPA自动通过", "已通过", "已通过"],
            weights=[0.4, 0.5, 0.1]
        )[0]

def build_base(prefix, category):
    """返回公共基础字段"""
    person, title = rand_person()
    dept = rand_dept()
    date = rand_date()
    risk_type, (anomaly_desc, score, risk_level) = pick_risk_pattern(category)
    status = derive_status(risk_level)
    project = random.choice(PROJECTS) if random.random() < 0.55 else "公用经费"
    return {
        "报销单号": rand_id(prefix),
        "申请日期": date.strftime("%Y-%m-%d"),
        "申请人": person,
        "职称": title,
        "所属部门": dept,
        "所属项目": project,
        "审批人": random.choice(APPROVERS),
        "银行": random.choice(BANKS),
        "审批状态": status,
        "风险等级": risk_level,
        "风险评分": score,
        "风险类型": risk_type,
        "异常行为描述": anomaly_desc,
    }, date

# ─── 各类别生成函数 ──────────────────────────────────────────

def gen_daily_expenses(n=200):
    subtypes = {
        "办公费":  ["文具耗材采购", "打印纸采购", "办公用品补充", "硒鼓墨盒采购"],
        "印刷费":  ["教材印刷", "资料装订", "论文打印", "宣传册印制", "试卷印刷"],
        "邮电费":  ["快递费", "邮寄材料费", "国际快递", "公文邮寄"],
        "市内交通费": ["公务打车", "地铁乘坐", "公交出行", "短途公务用车"],
        "维修费":  ["空调维修", "电脑维修", "实验室设备维修", "办公桌椅维修"],
        "材料费":  ["实验耗材", "化学试剂", "实验室器材", "教学材料"],
        "劳务费":  ["兼职助研劳务", "数据录入劳务", "学生助教劳务", "临时工劳务"],
        "会议费":  ["院系会议茶水", "学术研讨会场租", "培训会议餐费", "年度工作会议"],
    }
    invoice_map = {
        "办公费":    ("增值税普通发票", "购物收据"),
        "印刷费":    ("印刷服务发票",),
        "邮电费":    ("快递单据", "邮政收据"),
        "市内交通费": ("出租车发票", "网约车行程单"),
        "维修费":    ("维修服务发票", "维修清单"),
        "材料费":    ("采购发票", "验收单"),
        "劳务费":    ("劳务协议", "个人所得税扣缴记录"),
        "会议费":    ("会议室租用发票", "餐饮发票"),
    }
    rows = []
    for _ in range(n):
        base, date = build_base("RB", "日常报销")
        st = random.choice(list(subtypes.keys()))
        desc = random.choice(subtypes[st])
        inv_count = random.randint(0, 4)
        amount = round(random.uniform(50, 8000), 2)
        # 高风险时金额可能虚高
        if base["风险等级"] == "高风险":
            amount = round(random.uniform(3000, 20000), 2)
        base.update({
            "费用类别": "日常报销",
            "费用子类": st,
            "报销事由": desc,
            "报销金额（元）": amount,
            "票据类型": random.choice(invoice_map[st]),
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows


def gen_travel_expenses(n=200):
    destinations = ["北京", "上海", "广州", "深圳", "成都", "武汉", "南京", "杭州",
                    "西安", "重庆", "长沙", "郑州", "合肥", "厦门", "青岛"]
    purposes = ["参加学术会议", "出席学科评审", "科研合作交流", "参加培训",
                "教学考察", "课题组会议", "赴外学习", "项目验收"]
    transport_types = ["高铁二等座", "高铁一等座", "飞机经济舱", "飞机商务舱", "汽车", "高铁硬座"]
    hotel_levels = ["经济型酒店", "商务酒店", "三星酒店", "四星酒店", "五星酒店"]

    rows = []
    for _ in range(n):
        base, start_date = build_base("CL", "差旅费")
        days = random.randint(1, 7)
        end_date = start_date + timedelta(days=days)
        dest = random.choice(destinations)
        transport = random.choice(transport_types)
        hotel = random.choice(hotel_levels)

        transport_cost = round(random.uniform(200, 3000), 2)
        hotel_cost = round(random.uniform(150, 800) * days, 2)
        meal_subsidy = round(days * random.uniform(80, 120), 2)
        local_transport = round(days * random.uniform(30, 80), 2)
        total = round(transport_cost + hotel_cost + meal_subsidy + local_transport, 2)
        inv_count = random.randint(1, 6)

        # 高风险：超标酒店/商务舱/夹带私人费用
        if base["风险等级"] == "高风险":
            hotel = random.choice(["五星酒店", "豪华套房"])
            transport = random.choice(["飞机商务舱", "飞机头等舱"])
            hotel_cost = round(random.uniform(1000, 2500) * days, 2)
            total = round(transport_cost + hotel_cost + meal_subsidy + local_transport, 2)
            inv_count = random.randint(1, 3)

        base.update({
            "出发日期": start_date.strftime("%Y-%m-%d"),
            "返回日期": end_date.strftime("%Y-%m-%d"),
            "出差目的地": dest,
            "出差事由": f"{dest}·{random.choice(purposes)}（{days}天）",
            "交通工具": transport,
            "住宿标准": hotel,
            "出差天数": days,
            "交通费（元）": transport_cost,
            "住宿费（元）": hotel_cost,
            "伙食补助（元）": meal_subsidy,
            "市内交通（元）": local_transport,
            "合计金额（元）": total,
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows


def gen_asset_purchases(n=150):
    fixed_assets = [
        ("笔记本电脑", 8000, 15000), ("台式电脑", 5000, 12000),
        ("打印机", 3000, 8000), ("服务器", 20000, 80000),
        ("投影仪", 5000, 15000), ("实验仪器", 10000, 120000),
        ("显微镜", 15000, 50000), ("示波器", 8000, 30000),
        ("复印机", 6000, 15000), ("空调", 3000, 8000),
    ]
    intangible_assets = [
        ("正版操作系统", 1500, 5000), ("设计软件（AutoCAD）", 5000, 20000),
        ("统计分析软件（SPSS）", 8000, 25000), ("文献数据库订阅", 30000, 100000),
        ("专利授权费", 5000, 50000), ("ERP系统授权", 20000, 80000),
    ]
    rows = []
    for i in range(n):
        base, date = build_base("ZC", "资产采购")
        is_fixed = i < n * 0.7
        if is_fixed:
            name, lo, hi = random.choice(fixed_assets)
            asset_type, docs = "固定资产", "采购合同、验收报告、资产入库单"
        else:
            name, lo, hi = random.choice(intangible_assets)
            asset_type, docs = "无形资产", "软件授权协议、发票、验收确认函"

        amount = round(random.uniform(lo, hi), 2)
        qty = random.randint(1, 5) if amount < 20000 else 1
        total = round(amount * qty, 2)
        inv_count = random.randint(1, 3)
        supplier = random.choice(["京东商城", "华为授权店", "联想直销", "天猫官方店", "线下采购", "政府采购平台"])

        # 高风险：签订虚假合同/围标
        if base["风险等级"] == "高风险":
            supplier = random.choice(["关联企业-华X科技", "关联企业-明X贸易", "壳公司-新X网络"])
            amount = round(random.uniform(hi * 0.9, hi * 1.5), 2)
            total = round(amount * qty, 2)
            inv_count = random.randint(0, 1)
            docs = "（合同晚于验收日期）" + docs

        base.update({
            "资产类型": asset_type,
            "资产名称": name,
            "单价（元）": amount,
            "采购数量": qty,
            "合计金额（元）": total,
            "供应商": supplier,
            "所需单据": docs,
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows


def gen_advance_payments(n=150):
    purposes = [
        "会务费押金", "出差预借款", "设备采购预付款", "活动经费预支",
        "学术会议注册费预借", "实验耗材紧急采购", "学生活动经费预支", "科研设备定金"
    ]
    rows = []
    for _ in range(n):
        base, borrow_date = build_base("JK", "借款暂付款")
        repay_days = random.randint(7, 90)
        expected_repay = borrow_date + timedelta(days=repay_days)
        actual_repay = expected_repay + timedelta(days=random.randint(-5, 30))
        purpose = random.choice(purposes)
        amount = round(random.uniform(500, 30000), 2)
        is_settled = random.random() < 0.65
        settled_amount = round(amount * random.uniform(0.8, 1.0), 2) if is_settled else 0
        balance = round(amount - settled_amount, 2)
        inv_count = random.randint(0, 3)
        status_text = "已核销" if (is_settled and balance == 0) else ("部分核销" if is_settled else "待核销")

        # 高风险：金额偏大、转入关联账户、长期未还
        if base["风险等级"] == "高风险":
            amount = round(random.uniform(20000, 80000), 2)
            is_settled = False
            settled_amount = 0
            balance = amount
            status_text = "待核销"
            inv_count = 0
            repay_days = random.randint(90, 365)
            if base["异常行为描述"] == "正常报销":
                base["异常行为描述"] = "大额借款长期未核销，疑似资金挪用"
            base["风险评分"] = 6

        base.update({
            "借款事由": purpose,
            "借款金额（元）": amount,
            "预计还款日": expected_repay.strftime("%Y-%m-%d"),
            "实际核销日": actual_repay.strftime("%Y-%m-%d") if is_settled else "未核销",
            "已核销金额（元）": settled_amount,
            "未核销余额（元）": balance,
            "核销状态": status_text,
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows


def gen_labor_fees(n=200):
    labor_types = {
        "专家咨询费":  (3000, 30000, "咨询协议、专家信息表、税务登记"),
        "讲座费":     (500, 5000, "讲座邀请函、出席记录"),
        "学生劳务费":  (200, 3000, "劳务协议、学生信息表"),
        "校外兼职劳务": (1000, 10000, "劳务合同、个人银行卡信息"),
        "翻译费":     (500, 8000, "翻译合同、交付成果"),
        "评审费":     (500, 3000, "评审邀请函、出席证明"),
        "监考劳务费":  (50, 300, "监考安排表"),
        "助研劳务费":  (500, 5000, "科研助研协议"),
    }
    tax_rates = {
        "内部员工": 0.0,
        "外部个人（低于800）": 0.0,
        "外部个人（800-4000）": 0.2,
        "外部个人（4000以上）": 0.28,
        "企业": 0.06
    }
    rows = []
    for _ in range(n):
        base, date = build_base("LW", "劳务费")
        lt, (lo, hi, doc_req) = random.choice(list(labor_types.items()))
        recipient = f"{'外聘' if random.random()>0.3 else '校内'}{random.choice(['教师', '专家', '学生', '研究员'])}"
        amount = round(random.uniform(lo, hi), 2)

        # 高风险：冒名领取、向家属发放
        if base["风险等级"] == "高风险":
            recipient = random.choice(["申请人家属", "虚构人员-张某某", "项目无关人员"])
            amount = round(random.uniform(10000, 50000), 2)
            doc_req = "（协议缺失或伪造）" + doc_req
            base["风险评分"] = max(base["风险评分"], 5)

        if amount < 800:
            tax_type = "外部个人（低于800）"
        elif amount < 4000:
            tax_type = "外部个人（800-4000）"
        else:
            tax_type = random.choice(["外部个人（4000以上）", "企业"])

        tax_rate = tax_rates[tax_type]
        tax_amount = round(amount * tax_rate, 2)
        actual_payment = round(amount - tax_amount, 2)
        inv_count = random.randint(0, 3)
        if base["风险等级"] == "高风险":
            inv_count = random.randint(0, 1)

        base.update({
            "劳务类型": lt,
            "受付款人身份": recipient,
            "劳务说明": f"{lt}相关工作",
            "应付金额（元）": amount,
            "税务类型": tax_type,
            "代扣税率": f"{tax_rate*100:.0f}%",
            "代扣税额（元）": tax_amount,
            "实付金额（元）": actual_payment,
            "所需单据": doc_req,
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows


def gen_other_expenses(n=300):
    other_types = {
        "招待费": {
            "descs": ["接待外校专家", "重要客户来访接待", "项目合作洽谈餐", "校庆接待"],
            "amount": (200, 6000),
            "docs": "接待审批单、餐饮发票、参加人员名单"
        },
        "培训费": {
            "descs": ["参加专业技能培训", "管理干部培训", "新员工岗前培训", "安全教育培训"],
            "amount": (500, 8000),
            "docs": "培训通知、培训发票、结业证书复印件"
        },
        "委托业务费": {
            "descs": ["委托第三方数据采集", "委托律师法律服务", "委托会计师事务所审计", "委托软件开发"],
            "amount": (2000, 50000),
            "docs": "合同协议、业务完成证明、验收报告、正规发票"
        },
        "论文版面费": {
            "descs": ["SCI期刊版面费", "核心期刊版面费", "普通期刊版面费", "国际会议论文费"],
            "amount": (500, 15000),
            "docs": "期刊录用通知、版面费发票（收据）、论文首页复印件"
        },
        "邮电通讯费": {
            "descs": ["手机话费报销", "网络服务费", "国际长途费", "视频会议服务费"],
            "amount": (50, 1000),
            "docs": "通话记录单、运营商发票"
        },
        "设备维修维护费": {
            "descs": ["服务器年度维保", "实验室仪器校准", "中央空调系统维护", "网络设备维修"],
            "amount": (500, 30000),
            "docs": "维保合同、维修完成报告、服务发票"
        },
    }
    rows = []
    for _ in range(n):
        base, date = build_base("QT", "其他费用")
        ot, config = random.choice(list(other_types.items()))
        desc = random.choice(config["descs"])
        lo, hi = config["amount"]
        amount = round(random.uniform(lo, hi), 2)
        inv_count = random.randint(1, 4)

        # 高风险：虚假合同/关联交易
        if base["风险等级"] == "高风险":
            amount = round(random.uniform(hi * 0.8, hi * 2.0), 2)
            inv_count = random.randint(0, 1)
            desc = "（疑似虚构）" + desc
            config["docs"] = "（合同/发票存疑）" + config["docs"]

        base.update({
            "费用类别": ot,
            "报销事由": desc,
            "报销金额（元）": amount,
            "所需单据": config["docs"],
            "票据数量": inv_count,
        })
        rows.append(base)
    return rows

# ─── Excel 写入 ──────────────────────────────────────────────
RISK_COLORS = {
    "高风险": "FFD7D7",
    "中风险": "FFF3CD",
    "低风险": "E8F5E9",
}
RISK_TYPE_COLORS = {
    "源头虚构": "C00000",
    "金额操控": "E36C09",
    "支出失控": "7030A0",
    "流程违规": "1F497D",
    "正常":    "375623",
}

def style_header(cell, bg="1F4E79"):
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_data(cell, risk="低风险", risk_type="正常"):
    bg = RISK_COLORS.get(risk, "FFFFFF")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    color = RISK_TYPE_COLORS.get(risk_type, "000000") if risk != "低风险" else "000000"
    cell.font = Font(name="微软雅黑", size=9, color=color)

def write_sheet(wb, sheet_name, rows):
    if not rows:
        return
    ws = wb.create_sheet(title=sheet_name)
    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    for ri, row in enumerate(rows, 2):
        rv = row.get("风险等级", "低风险")
        rt = row.get("风险类型", "正常")
        for ci, key in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row[key])
            style_data(cell, rv, rt)

    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        max_len = max(len(str(h)), max((len(str(r.get(h, ""))) for r in rows), default=0))
        ws.column_dimensions[col_letter].width = min(max_len * 1.5 + 2, 36)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def write_summary_sheet(wb, category_map):
    from collections import Counter
    all_data = []
    for records in category_map.values():
        all_data.extend(records)

    ws = wb.create_sheet(title="汇总分析", index=0)
    ws.column_dimensions["A"].width = 24
    for col in ["B","C","D","E","F"]:
        ws.column_dimensions[col].width = 16

    def hdr(r, c, text, bg="1F4E79"):
        cell = ws.cell(row=r, column=c, value=text)
        style_header(cell, bg)

    def val(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="微软雅黑", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "高校财务报销模拟数据集 v2 — 风险分析汇总"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 34

    ws["A2"] = f"生成日期：{datetime.now().strftime('%Y-%m-%d')}    总记录数：{len(all_data)} 条"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="595959")
    ws.merge_cells("A2:F2")

    # ── 各类别风险分布表 ──
    row = 4
    for h, c in [("费用类别","A"),("记录数","B"),("低风险","C"),("中风险","D"),("高风险","E"),("高风险占比","F")]:
        hdr(row, "ABCDEF".index(c)+1, h)
    row += 1

    for cat, records in category_map.items():
        low  = sum(1 for r in records if r.get("风险等级") == "低风险")
        mid  = sum(1 for r in records if r.get("风险等级") == "中风险")
        high = sum(1 for r in records if r.get("风险等级") == "高风险")
        val(row, 1, cat)
        val(row, 2, len(records))
        val(row, 3, low)
        val(row, 4, mid)
        val(row, 5, high)
        val(row, 6, f"{high/len(records)*100:.1f}%" if records else "0%")
        # 高风险行红色背景
        if high / len(records) > 0.15:
            ws.cell(row, 6).fill = PatternFill("solid", start_color="FFD7D7")
        row += 1

    # ── 风险类型分布 ──
    row += 2
    hdr(row, 1, "风险类型")
    hdr(row, 2, "记录数")
    hdr(row, 3, "占比")
    hdr(row, 4, "主要危害", bg="7B2D2D")
    ws.merge_cells(f"D{row}:F{row}")
    row += 1
    risk_type_info = {
        "源头虚构": "虚假票据/事项/合同套取资金，性质最严重",
        "金额操控": "拆单规避审批、转移资金至关联方",
        "支出失控": "挪用专项资金、超范围支出、夹报私费",
        "流程违规": "先采购后补批、合同倒签、审批缺位",
        "正常":    "无异常",
    }
    cnt_by_type = Counter(r.get("风险类型","正常") for r in all_data)
    for rt in ["源头虚构","金额操控","支出失控","流程违规","正常"]:
        cnt = cnt_by_type.get(rt, 0)
        val(row, 1, rt)
        val(row, 2, cnt)
        val(row, 3, f"{cnt/len(all_data)*100:.1f}%")
        cell = ws.cell(row=row, column=4, value=risk_type_info[rt])
        ws.merge_cells(f"D{row}:F{row}")
        cell.font = Font(name="微软雅黑", size=9, color=RISK_TYPE_COLORS.get(rt,"000000"))
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row += 1

    # ── 审批状态分布 ──
    row += 2
    hdr(row, 1, "审批状态")
    hdr(row, 2, "记录数")
    hdr(row, 3, "占比")
    row += 1
    for s, cnt in Counter(r.get("审批状态","") for r in all_data).most_common():
        if s:
            val(row, 1, s)
            val(row, 2, cnt)
            val(row, 3, f"{cnt/len(all_data)*100:.1f}%")
            row += 1

    ws.freeze_panes = "A3"


# ─── 主函数 ──────────────────────────────────────────────────
def main():
    print("正在生成模拟报销数据 v3（5200条记录）...")

    category_map = {
        "日常报销":   gen_daily_expenses(867),
        "差旅费":     gen_travel_expenses(867),
        "资产采购":   gen_asset_purchases(650),
        "借款暂付款": gen_advance_payments(650),
        "劳务费":     gen_labor_fees(867),
        "其他费用":   gen_other_expenses(1299),
    }
    all_data = []
    for v in category_map.values():
        all_data.extend(v)
    print(f"共生成 {len(all_data)} 条记录")

    # 打印风险分布统计
    from collections import Counter
    print("\n=== 风险等级分布 ===")
    for cat, records in category_map.items():
        cnt = Counter(r.get("风险等级","") for r in records)
        print(f"  {cat}: 低={cnt['低风险']} 中={cnt['中风险']} 高={cnt['高风险']}")
    print("\n=== 风险类型分布 ===")
    print(Counter(r.get("风险类型","") for r in all_data))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, category_map)
    for sheet_name, rows in category_map.items():
        write_sheet(wb, sheet_name, rows)

    output_path = r"C:\Users\a120\WorkBuddy\20260422055514\高校财务报销模拟数据集_v3_5200条.xlsx"
    wb.save(output_path)
    print(f"\n已保存至：{output_path}")

if __name__ == "__main__":
    main()
